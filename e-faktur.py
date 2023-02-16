import os.path
import csv
import sys
import openpyxl
import requests
import pyautogui
import cv2
import shutil as sh
import numpy as np
import glob
# from pyzbar.pyzbar import decode
from PIL import Image
from pdf2image import convert_from_path
import xml.etree.ElementTree as eTree

message = pyautogui.confirm('Apakah program ingin dijalankan?', buttons=['OK','Cancel'])
for x in message:
    try:
        if message == 'OK':
            continue
        else:
            sys.exit()
    except:
        sys.exit()
        
# def barcode_reader(filename):
#     try:
#         img = cv2.imread(filename)
#         qr_code = decode(img)
#         for code in qr_code:
#             return code.data.decode("utf-8")
#     except NameError:
#         print('Failed') 

def barcode_reader(filename):
    try:
        img = cv2.imread(filename)
        detect = cv2.QRCodeDetector()
        value, points, straight_qrcode = detect.detectAndDecode(img)
        return value
    except NameError:
        return "Failed"

base_dir = "."

disinipdf = (os.path.join(base_dir, "disini_pdf"))
disinijpg = (os.path.join(base_dir, "disini_image"))
disinixls = (os.path.join(base_dir, "disini_excel"))
disiniNG = (os.path.join(base_dir, "disini_NG"))

try:
    os.mkdir(disinipdf)
    os.mkdir(disinijpg)
    os.mkdir(disinixls)
    os.mkdir(disiniNG)
except:
    print("Direktori sudah ada")

no_page = 0
isiQR = []
daftar_input = os.listdir(disinipdf)
print(daftar_input)
for filepdf in daftar_input:
    gambarf_pdf = convert_from_path(disinipdf + "/" + filepdf, dpi=500)
    for i in range(len(gambarf_pdf)):
        no_page += 1
        gambarf_pdf[i].save(f'{disinijpg}/page{no_page}.jpg', 'JPEG')
        try:
                
            link = barcode_reader(f'{disinijpg}/page{no_page}.jpg')
            if link == '':
                isiQR.append('')
                continue
            if link not in isiQR:
                isiQR.append(link)
            
        except:
            continue 

print(isiQR)

no_link = 0
isi_csv = []
for qr in isiQR:
    no_link += 1      
    if qr == '':
        continue
    while True:
        try:
            resp = requests.get(qr, timeout=1)
            print(resp.content.decode("utf-8").__contains__("No service was found."))
            if resp.content.decode("utf-8").__contains__("No service was found."):
                # simpan ke sqlite dan aktifkan worker, GAJADI, DIA PENDING DOWNLOAD KECUALI KALO ADA KONEKSI
                pass
            else:
                break
        except:
            break
    print(resp.content)

    # saving the xml file
    nama_xml = f'{disinijpg}/link{no_link}.xml'
    with open(nama_xml, 'wb') as f:
        f.write(resp.content)

    tree = eTree.parse(nama_xml)
    root = tree.getroot()
    key_data = {}
    print(root.tag)
    for child in root:
        print(child.tag)
        if child.tag == "detailTransaksi":
            for grandChild in child:
                print(grandChild.tag)
                key_data[grandChild.tag] = grandChild.text
            break

        key_data[child.tag] = child.text

    print(key_data)

    # save ke folder NG
    rs = r"D:\allprojectrandom\e-faktur-auto\disini_image"
    ds = r"D:\allprojectrandom\e-faktur-auto\disini_NG"
    
    pattern = '\*.jpg'
    files = glob.glob(rs + pattern)
    
    try:
        for file in files:
            if link == '':
                isiQR.append('')
                continue
            if link not in isiQR:
                isiQR.append(link)
        file_names = os.path.basename(file)
        sh.move(file, rs + file_names)
    except NameError:
        print("Gagal")


    nama_xls = f'{disinixls}/cek_lalu_upload.xlsx'
    if os.path.exists(nama_xls):
        workbook = openpyxl.load_workbook(filename=nama_xls)
    else:
        workbook = openpyxl.Workbook()
    worksheet = workbook.active
    kols = {
        'NoFaktur': "A",
        'KdJenisTransaksi': "B",
        'fgPengganti': "C",
        'NoFakturAsli': "D",
        'Date': "E",
        'NamaPenjual': "F",
        'AlamatPenjual': "G",
        'NPWPPenjual': "H",
        'NamaPembeli': "I",
        'AlamatPembeli': "J",
        'NPWPPembeli': "K",
        'dpp': "L",
        'ppn': "M",
        'ppnbm': "N",
        'statusApproval': "O",
        'statusFaktur': "P",
    }
    fields = {
        'NoFaktur': "gabungan",
        'KdJenisTransaksi': "kdJenisTransaksi",
        'fgPengganti': "fgPengganti",
        'NoFakturAsli': "nomorFaktur",
        'Date': "tanggalFaktur",
        'NamaPenjual': "namaPenjual",
        'AlamatPenjual': "alamatPenjual",
        'NPWPPenjual': "npwpPenjual",
        'NamaPembeli': "namaLawanTransaksi",
        'AlamatPembeli': "alamatLawanTransaksi",
        'NPWPPembeli': "npwpLawanTransaksi",
        'dpp': "dpp",
        'ppn': "ppn",
        'ppnbm': "ppnbm",
        'statusApproval': "statusApproval",
        'statusFaktur': "statusFaktur",
    }
    for key, value in kols.items():
        worksheet[f'{value}1'] = key

    header_csv = [
        'FM',
        'KD_JENIS_TRANSAKSI',
        'FG_PENGGANTI',
        'NOMOR_FAKTUR',
        'MASA_PAJAK',
        'TAHUN_PAJAK',
        'TANGGAL_FAKTUR',
        'NPWP',
        'NAMA',
        'ALAMAT_LENGKAP',
        'JUMLAH_DPP',
        'JUMLAH_PPN',
        'JUMLAH_PPNBM',
        'IS_CREDITABLE',
    ]
    rows = worksheet.max_row
    rows += 1
    for key, value in fields.items():

        breakflag = False
        entri_csv = []
        indeks = f'{kols[key]}{rows}'
    
        if value == "gabungan":
            try:
                no_faktur = f'{key_data["kdJenisTransaksi"]}{key_data["fgPengganti"]}.{key_data["nomorFaktur"]}'
            except:
                continue
            for row in worksheet.iter_rows(min_row=1, max_row=rows, min_col=0, max_col=0, values_only=False):
                for cell in row:
                    if cell.value == no_faktur:
                        print(f"{cell.value} sama dengan {no_faktur}")

                        breakflag = True
                        break
            if breakflag:
                break
            else:
                worksheet[indeks] = no_faktur
                continue
        worksheet[indeks] = key_data[value]
    workbook.save(filename=nama_xls)

    pecahan_tanggal = key_data['tanggalFaktur'].split('/')
    if key_data['kdJenisTransaksi'] in ['04','05','07','08']:
        is_creditable = 0
    else:
        is_creditable = 1

    entri_csv.append('FM')
    entri_csv.append(key_data['kdJenisTransaksi'])
    entri_csv.append(key_data['fgPengganti'])
    entri_csv.append(key_data['nomorFaktur'])
    entri_csv.append(pecahan_tanggal[1])
    entri_csv.append(pecahan_tanggal[-1])
    entri_csv.append(key_data['tanggalFaktur'])
    entri_csv.append(key_data['npwpPenjual'])
    entri_csv.append(key_data['namaPenjual'])
    entri_csv.append(key_data['alamatPenjual'])
    entri_csv.append(key_data['jumlahDpp'])
    entri_csv.append(key_data['jumlahPpn'])
    entri_csv.append(key_data['jumlahPpnBm'])
    entri_csv.append(is_creditable)
    isi_csv.append(entri_csv)

    print(entri_csv)
    print(isi_csv)

with open(f'{disinixls}/cek_csv.csv', 'w', encoding='UTF8', newline="") as f:
    writer = csv.writer(f, delimiter=";")
    writer.writerow(header_csv)
    writer.writerows(isi_csv)

# save untuk daftar xml
nama_link = f'{disinixls}/daftar_link.xlsx'
if os.path.exists(nama_link):
    workbook = openpyxl.load_workbook(filename=nama_link)
else:
    workbook = openpyxl.Workbook()
worksheet = workbook.active
rows = 0
for qr in isiQR:
    rows += 1
    worksheet[f'A{rows}'] = qr
workbook.save(filename=nama_link)
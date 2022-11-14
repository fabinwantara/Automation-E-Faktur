import os.path
import openpyxl
import csv
import cv2
import requests
from pdf2image import convert_from_path
import xml.etree.ElementTree as eTree


def read_qr_code(filename):
    try:
        img = cv2.imread(filename)
        detect = cv2.QRCodeDetector()
        value, points, straight_qrcode = detect.detectAndDecode(img)
        return value
    except NameError:
        return "gagal"


# base_dir = os.path.join(os.environ["HOMEPATH"], "Desktop")
base_dir = "."
# base_dir = "C:/Users/DESK45/Documents"
# if __file__.endswith('.exe'):
#     base_dir = os.path.join(os.path.dirname(__file__), 'foldername')
#     if not os.path.exists(base_dir):
#         os.makedirs(base_dir)
disinipdf = (os.path.join(base_dir, "disini_pdf"))
disinijpg = (os.path.join(base_dir, "disini_image"))
disinixls = (os.path.join(base_dir, "disini_excel"))
try:
    os.mkdir(disinipdf)
    os.mkdir(disinijpg)
    os.mkdir(disinixls)
except:
    print("sudah ada direktori")

no_page = 0
isiQR = []
daftar_input = os.listdir(disinipdf)
print(daftar_input)
for filepdf in daftar_input:
    gambarf_pdf = convert_from_path(disinipdf + "/" + filepdf)
    for i in range(len(gambarf_pdf)):
        gambarf_pdf[i].save(f'{disinijpg}/page{no_page}.jpg', 'JPEG')
        link = read_qr_code(f'{disinijpg}/page{no_page}.jpg')
        if link == '':
            continue
        isiQR.append(link)
        no_page += 1


print(isiQR)
no_link = 0
isi_csv = []
for qr in isiQR:
    while True:
        try:
            resp = requests.get(qr, timeout=1)
            print(resp.content.decode("utf-8").__contains__("No service was found."))
            if resp.content.decode("utf-8").__contains__("No service was found."):
                # simpan ke sqlite dan aktifkan worker, GAJADI, DIA PENDING DOWNLOAD KECUALI KALO ADA KONEKSI
                pass
            else:
                break
        except:
            break
    print(resp.content)

    # saving the xml file
    nama_xml = f'{disinijpg}/link{no_link}.xml'
    with open(nama_xml, 'wb') as f:
        f.write(resp.content)
        no_link += 1

    tree = eTree.parse(nama_xml)
    root = tree.getroot()
    key_data = {}
    print(root.tag)
    for child in root:
        print(child.tag)
        if child.tag == "detailTransaksi":
            for grandChild in child:
                print(grandChild.tag)
                key_data[grandChild.tag] = grandChild.text
            break

        key_data[child.tag] = child.text

    print(key_data)

    nama_xls = f'{disinixls}/cek_lalu_upload.xlsx'
    if os.path.exists(nama_xls):
        wookbook = openpyxl.load_workbook(filename=nama_xls)
    else:
        wookbook = openpyxl.Workbook()
    worksheet = wookbook.active
    kols = {
        'NoFaktur': "A",
        'KdJenisTransaksi': "B",
        'fgPengganti': "C",
        'NoFakturAsli': "D",
        'Date': "E",
        'NamaPenjual': "F",
        'AlamatPenjual': "G",
        'NPWPPenjual': "H",
        'NamaPembeli': "I",
        'AlamatPembeli': "J",
        'NPWPPembeli': "K",
        'dpp': "L",
        'ppn': "M",
        'ppnbm': "N",
        'statusApproval': "O",
        'statusFaktur': "P",
    }
    fields = {
        'NoFaktur': "gabungan",
        'KdJenisTransaksi': "kdJenisTransaksi",
        'fgPengganti': "fgPengganti",
        'NoFakturAsli': "nomorFaktur",
        'Date': "tanggalFaktur",
        'NamaPenjual': "namaPenjual",
        'AlamatPenjual': "alamatPenjual",
        'NPWPPenjual': "npwpPenjual",
        'NamaPembeli': "namaLawanTransaksi",
        'AlamatPembeli': "alamatLawanTransaksi",
        'NPWPPembeli': "npwpLawanTransaksi",
        'dpp': "dpp",
        'ppn': "ppn",
        'ppnbm': "ppnbm",
        'statusApproval': "statusApproval",
        'statusFaktur': "statusFaktur",
    }
    for key, value in kols.items():
        worksheet[f'{value}1'] = key

    header_csv = [
        'FM',
        'KD_JENIS_TRANSAKSI',
        'FG_PENGGANTI',
        'NOMOR_FAKTUR',
        'MASA_PAJAK',
        'TAHUN_PAJAK',
        'TANGGAL_FAKTUR',
        'NPWP',
        'NAMA',
        'ALAMAT_LENGKAP',
        'JUMLAH_DPP',
        'JUMLAH_PPN',
        'JUMLAH_PPNBM',
        'IS_CREDITABLE',
    ]
    rows = worksheet.max_row
    rows += 1
    for key, value in fields.items():
        # rows += 1
        breakflag = False
        entri_csv = []
        indeks = f'{kols[key]}{rows}'
        # try:
        if value == "gabungan":
            no_faktur = f'{key_data["kdJenisTransaksi"]}{key_data["fgPengganti"]}.{key_data["nomorFaktur"]}'
            for row in worksheet.iter_rows(min_row=1, max_row=rows, min_col=0, max_col=0, values_only=False):
                for cell in row:
                    if cell.value == no_faktur:
                        print(f"{cell.value} sama dengan {no_faktur}")
                        # raise StopIteration
                        breakflag = True
                        break
                # if breakflag:
                #     break
            if breakflag:
                break
            else:
                worksheet[indeks] = no_faktur
                continue
        worksheet[indeks] = key_data[value]
    wookbook.save(filename=nama_xls)

    pecahan_tanggal = key_data['tanggalFaktur'].split('/')
    if key_data['kdJenisTransaksi'] == '07':
        is_creditable = 0
    else:
        is_creditable = 1

    entri_csv.append('FM')
    entri_csv.append(key_data['kdJenisTransaksi'])
    entri_csv.append(key_data['fgPengganti'])
    entri_csv.append(key_data['nomorFaktur'])
    entri_csv.append(pecahan_tanggal[1])
    entri_csv.append(pecahan_tanggal[-1])
    entri_csv.append(key_data['tanggalFaktur'])
    entri_csv.append(key_data['npwpPenjual'])
    entri_csv.append(key_data['namaPenjual'])
    entri_csv.append(key_data['alamatPenjual'])
    entri_csv.append(key_data['jumlahDpp'])
    entri_csv.append(key_data['jumlahPpn'])
    entri_csv.append(key_data['jumlahPpnBm'])
    entri_csv.append(is_creditable)
    isi_csv.append(entri_csv)

    print(entri_csv)
    print(isi_csv)

with open(f'{disinixls}/cek_csv.csv', 'w', encoding='UTF8', newline="") as f:
    writer = csv.writer(f, delimiter=";")
    writer.writerow(header_csv)
    writer.writerows(isi_csv)

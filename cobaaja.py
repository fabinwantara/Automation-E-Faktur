import os.path
import time
import numpy as np
# import PyPDF2 as PyPDF2
# from pyzbar import pyzbar
import openpyxl
import csv
import cv2
import requests
from pdf2image import convert_from_path
import xml.etree.ElementTree as eTree


# def display(im, bbox):
#     n = len(bbox)
#     for j in range(n):
#         cv2.line(im, tuple(bbox[j][0]), tuple(bbox[(j + 1) % n][0]), (255, 0, 0), 3)
#
#     # Display results
#     cv2.imshow("Results", im)

def read_qr_code(filename):
    try:
        img = cv2.imread(filename)
        detect = cv2.QRCodeDetector()
        value2, points, straight_qrcode = detect.detectAndDecode(img)
        # if len(value) > 0:
        #     print("Decoded Data : {}".format(value))
        #     display(img, points)
        #     rectifiedImage = np.uint8(straight_qrcode)
        #     cv2.imshow("Rectified QRCode", rectifiedImage)
        # else:
        #     print("QR Code not detected")
        #     cv2.imshow("Results", img)
        return value2
    except NameError:
        return "gagal"


# base_dir = os.path.join(os.environ["HOMEPATH"], "Desktop")
base_dir = "."
# base_dir = "C:/Users/DESK45/Documents"
# if _file_.endswith('.exe'):
#     base_dir = os.path.join(os.path.dirname(_file_), 'foldername')
#     if not os.path.exists(base_dir):
#         os.makedirs(base_dir)
disinipdf2 = (os.path.join(base_dir, "disini_pdf2"))
disinipdf = (os.path.join(base_dir, "disini_pdf"))
disinijpg = (os.path.join(base_dir, "disini_image"))
disinixls = (os.path.join(base_dir, "disini_excel"))
try:
    os.mkdir(disinipdf)
    os.mkdir(disinijpg)
    os.mkdir(disinixls)
except:
    print("sudah ada direktori")

no_page = 0
isiQR = []

# daftar_input2 = os.listdir(disinipdf2)
# reader = PyPDF2.PdfReader(f"{disinipdf2}/05122022112317.pdf")
# for page in range(len(reader.pages)):
#     writer = PyPDF2.PdfWriter()
#     writer.addPage(reader.getPage(page))
#     with open(f'{disinipdf}/output{page+1}.pdf', 'wb') as outfile:
#         writer.write(outfile)

daftar_input = os.listdir(disinipdf)
print(daftar_input)
for filepdf in daftar_input:
    gambarf_pdf = convert_from_path(disinipdf + "/" + filepdf, dpi=700)
    for i in range(len(gambarf_pdf)):
        no_page += 1
        # if no_page not in [1, 2, 3, 4, 5, 7]:
        #     continue
        gambarf_pdf[i].save(f'{disinijpg}/page{no_page}.jpg', 'JPEG')
        # imggg = cv2.imread(f'{disinijpg}/page{no_page}.jpg', cv2.IMREAD_GRAYSCALE)
        # # im_gray = cv2.cvtColor(imggg, cv2.COLOR_BGR2GRAY)
        # # im_blur = cv2.GaussianBlur(imggg, (9, 9), 0)
        # im_thresh = cv2.threshold(imggg, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]
        #
        # (h, w) = im_thresh.shape[:2]
        # center = (w / 2, h / 2)
        # angle = 0
        # # angle = 180
        # scale = 0.5
        #
        # M = cv2.getRotationMatrix2D(center, angle, scale)
        # im_rotate = cv2.warpAffine(im_thresh, M, (w, h))
        #
        # # Morph close
        # kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (5, 5))
        # close = cv2.morphologyEx(im_rotate, cv2.MORPH_CLOSE, kernel, iterations=2)
        #
        # # Find contours and filter for QR code
        # cnts = cv2.findContours(close, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        # cnts = cnts[0] if len(cnts) == 2 else cnts[1]
        #
        # for c in cnts:
        #     peri = cv2.arcLength(c, True)
        #     approx = cv2.approxPolyDP(c, 0.04 * peri, True)
        #     x, y, w, h = cv2.boundingRect(approx)
        #     area = cv2.contourArea(c)
        #     ar = w / float(h)
        #     if len(approx) == 4 and area > 1000 and (ar > .85 and ar < 1.3):
        #         cv2.rectangle(im_rotate, (x, y), (x + w, y + h), (36, 255, 12), 1)
        #         ROI = im_rotate[y-10:y + h+10, x-10:x + w+10]
        #         cv2.imwrite(f'{disinijpg}/qr{no_page}.jpg', cv2.bitwise_not(ROI))

        # cv2.imwrite(f'{disinijpg}/page{no_page}.jpg', im_rotate)
        # cv2.imshow('thresh', im_thresh)
        # cv2.imshow('close', close)
        # cv2.imshow('image', imggg)
        # cv2.imshow('ROI', ROI)
        # cv2.waitKey()
        try:
            # image = cv2.imread(f'{disinijpg}/page{no_page}.jpg')
            # image = cv2.imread(f'{disinijpg}/qr{no_page}.jpg')
            # barcodes = pyzbar.decode(image)
            # for barcode in barcodes:
            #     (x, y, w, h) = barcode.rect
            #     cv2.rectangle(image, (x, y), (x + w, y + h), (0, 0, 255), 2)
            #
            #     # the barcode data is a bytes object so if we want to draw it on
            #     # our output image we need to convert it to a string first
            #     barcodeData = barcode.data.decode("utf-8")
            #     barcodeType = barcode.type

            link = read_qr_code(f'{disinijpg}/page{no_page}.jpg')
            # link = read_qr_code(f'{disinijpg}/qr{no_page}.jpg')
            if link == '':
                # isiQR.append(barcodeData)
                isiQR.append('')
                continue
            if link not in isiQR:
                isiQR.append(link)
        except:
            continue


print(isiQR)
no_link = 0
isi_csv = []
for qr in isiQR:
    no_link += 1
    if qr == '':
        continue
    while True:
        try:
            resp = requests.get(qr, timeout=1)
            print(resp.content.decode("utf-8")._contains_("No service was found."))
            if resp.content.decode("utf-8")._contains_("No service was found."):
                # simpan ke sqlite dan aktifkan worker, GAJADI, DIA PENDING DOWNLOAD KECUALI KALO ADA KONEKSI
                pass
            else:
                break
        except:
            break
    print(resp.content)

    # saving the xml file
    nama_xml = f'{disinijpg}/link{no_link}.xml'
    with open(nama_xml, 'wb') as f:
        f.write(resp.content)
        # no_link += 1

    tree = eTree.parse(nama_xml)
    root = tree.getroot()
    key_data = {}
    print(root.tag)
    for child in root:
        print(child.tag)
        if child.tag == "detailTransaksi":
            for grandChild in child:
                print(grandChild.tag)
                key_data[grandChild.tag] = grandChild.text
            break

        key_data[child.tag] = child.text

    print(key_data)

    nama_xls = f'{disinixls}/cek_lalu_upload.xlsx'
    if os.path.exists(nama_xls):
        wookbook = openpyxl.load_workbook(filename=nama_xls)
    else:
        wookbook = openpyxl.Workbook()
    worksheet = wookbook.active
    kols = {
        'NoFaktur': "A",
        'KdJenisTransaksi': "B",
        'fgPengganti': "C",
        'NoFakturAsli': "D",
        'Date': "E",
        'NamaPenjual': "F",
        'AlamatPenjual': "G",
        'NPWPPenjual': "H",
        'NamaPembeli': "I",
        'AlamatPembeli': "J",
        'NPWPPembeli': "K",
        'dpp': "L",
        'ppn': "M",
        'ppnbm': "N",
        'statusApproval': "O",
        'statusFaktur': "P",
    }
    fields = {
        'NoFaktur': "gabungan",
        'KdJenisTransaksi': "kdJenisTransaksi",
        'fgPengganti': "fgPengganti",
        'NoFakturAsli': "nomorFaktur",
        'Date': "tanggalFaktur",
        'NamaPenjual': "namaPenjual",
        'AlamatPenjual': "alamatPenjual",
        'NPWPPenjual': "npwpPenjual",
        'NamaPembeli': "namaLawanTransaksi",
        'AlamatPembeli': "alamatLawanTransaksi",
        'NPWPPembeli': "npwpLawanTransaksi",
        'dpp': "dpp",
        'ppn': "ppn",
        'ppnbm': "ppnbm",
        'statusApproval': "statusApproval",
        'statusFaktur': "statusFaktur",
    }
    for key, value in kols.items():
        worksheet[f'{value}1'] = key

    header_csv = [
        'FM',
        'KD_JENIS_TRANSAKSI',
        'FG_PENGGANTI',
        'NOMOR_FAKTUR',
        'MASA_PAJAK',
        'TAHUN_PAJAK',
        'TANGGAL_FAKTUR',
        'NPWP',
        'NAMA',
        'ALAMAT_LENGKAP',
        'JUMLAH_PPN',
        'JUMLAH_PPNBM',
        'IS_CREDITABLE',
    ]
    rows = worksheet.max_row
    rows += 1
    for key, value in fields.items():
        # rows += 1
        breakflag = False
        entri_csv = []
        indeks = f'{kols[key]}{rows}'
        # try:
        if value == "gabungan":
            try:
                no_faktur = f'{key_data["kdJenisTransaksi"]}{key_data["fgPengganti"]}.{key_data["nomorFaktur"]}'
            except:
                continue
            for row in worksheet.iter_rows(min_row=1, max_row=rows, min_col=0, max_col=0, values_only=False):
                for cell in row:
                    if cell.value == no_faktur:
                        print(f"{cell.value} sama dengan {no_faktur}")
                        # raise StopIteration
                        breakflag = True
                        break
                # if breakflag:
                #     break
            if breakflag:
                break
            else:
                worksheet[indeks] = no_faktur
                continue
        worksheet[indeks] = key_data[value]
    wookbook.save(filename=nama_xls)

    pecahan_tanggal = key_data['tanggalFaktur'].split('/')
    if key_data['kdJenisTransaksi'] in ['04', '05', '07', '08']:
        is_creditable = 0
    else:
        is_creditable = 1

    entri_csv.append('FM')
    entri_csv.append(key_data['kdJenisTransaksi'])
    entri_csv.append(key_data['fgPengganti'])
    entri_csv.append(key_data['nomorFaktur'])
    entri_csv.append(pecahan_tanggal[1])
    entri_csv.append(pecahan_tanggal[-1])
    entri_csv.append(key_data['tanggalFaktur'])
    entri_csv.append(key_data['npwpPenjual'])
    entri_csv.append(key_data['namaPenjual'])
    entri_csv.append(key_data['alamatPenjual'])
    entri_csv.append(key_data['jumlahDpp'])
    entri_csv.append(key_data['jumlahPpn'])
    entri_csv.append(key_data['jumlahPpnBm'])
    entri_csv.append(is_creditable)
    if entri_csv not in isi_csv:
        isi_csv.append(entri_csv)

    print(entri_csv)
    print(isi_csv)

with open(f'{disinixls}/cek_csv.csv', 'w', encoding='UTF8', newline="") as f:
    writer = csv.writer(f, delimiter=";")
    writer.writerow(header_csv)
    writer.writerows(isi_csv)

nama_xls2 = f'{disinixls}/daftar_link.xlsx'
if os.path.exists(nama_xls2):
    wookbook = openpyxl.load_workbook(filename=nama_xls2)
else:
    wookbook = openpyxl.Workbook()
worksheet = wookbook.active
rows = 0
for qr in isiQR:
    rows += 1
    worksheet[f'A{rows}'] = qr
wookbook.save(filename=nama_xls2)